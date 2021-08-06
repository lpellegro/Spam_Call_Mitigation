import requests
import json
import os
from csv import writer
from datetime import datetime
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

from requests_toolbelt.multipart.encoder import MultipartEncoder
from requests.auth import HTTPBasicAuth
from webex_bot.models.command import Command
from webex_bot.cards.echo_card import ECHO_CARD_CONTENT
from webex_bot.formatting import quote_info
from credentials import credentials

#THIS SCRIPT WORKS TOGETHER WITH WEBEX_BOT. WHEN "MORE INFO" IS PUSHED, IT RETURNS THE CALL ACTIVITY LIST

def cdrcheck(expe, username, secret, newip):
    
    new_url='https://' + expe + '/api/management/status/call/call'
    print('Sending query to: ', new_url)
    response = requests.get(new_url, auth=HTTPBasicAuth(username, secret))
    print("Query sent to:", new_url)
    #print(response)
    r=response.json()
    #print ('La risposta è', r)
    print('La lunghezza della risposta è:', len(r))
    if len(r)==0:   
       print("JSON empty answer")
       return
    peers=len(r)
    source_alias=''
    destination_alias=''
    day=''
    time=''
    nameisp=''
    countryisp=''
    emailisp=''
    apikey = credentials['apikey']
    url_whois = 'https://ip-netblocks.whoisxmlapi.com/api/v2?apiKey='+apikey+'&ip='+ newip
    
    try:
       response_whois = requests.get(url_whois)
       r_whois = response_whois.json()
       if 'PRIVATE-ADDRESS' in r_whois['result']['inetnums'][0]['netname']:
          nameisp='PRIVATE ADDRESS'
       else:
          nameisp=r_whois['result']['inetnums'][0]['as']['name']
          
       if nameisp == '':
          nameisp = 'Unknown'
   
       email_list=r_whois['result']['inetnums'][0]['abuseContact']
       if not email_list:
          print ('Empty email')
          emailisp = 'Unknown'
       else:
         emailisp=email_list[0]
         if len (emailisp) > 0:
            emailisp = 'Unknown'
   
       countryisp = r_whois['result']['inetnums'][0]['country']
       if countryisp == '':
          countryisp = 'Unknown'
     
    except requests.exceptions.ConnectionError:
       nameisp = 'Unknown'
       emailisp ='Unknown'
       countryisp = 'Unknown'
   
    print (nameisp, emailisp, countryisp)
    
    data=[]
    data.append(['IP Address:', 'Owner:', 'Country:', 'Email address:'])
    data.append([newip, nameisp, countryisp, emailisp])
    data.append(["Calling", "Called", "Timestamp", "Disconnect Reason"]) 
    for peer in range(peers):
        print ('Current peer index: ', peer)
        print ('Peer cdr: ', r[peer])
        calls=r[peer]['num_recs']
        #print("Chiamate", calls)
        call_record=r[peer]['records']
        #print("call_record", call_record)
        for call in range(calls):
            #print ('numero chiamata', call)
            call_details=call_record[call]['details'] #string just works fine
            if newip in call_details:
               disconnect_reason=call_record[call]['disconnect_reason']
               source_alias=call_record[call]['source_alias']
               destination_alias=call_record[call]['destination_alias']
               timestamp=call_record[call]['start_time'][:-7]
               timestamp_list=timestamp.split(" ")
               day=timestamp_list[0]
               time=timestamp_list[1]
               #protocol=call_record[call]['protocol'][:-4]
               data.append([source_alias, destination_alias, timestamp, disconnect_reason])
              
    print('Activity list in list:', data)           
    return data

class SendfileCommand(Command):

    def __init__(self):
        super().__init__(
            command_keyword="callback___sendfile",
            help_message="IP Info",
            card=ECHO_CARD_CONTENT)

    def execute(self, message, attachment_actions):
        """
        If you want to respond to a submit operation on the card, you
        would write code here!

        You can return text string here or even another card (Response).

        This sample command function simply echos back the sent message.

        :param message: message with command already stripped
        :param attachment_actions: attachment_actions object
        :return: a string or Response object. Use Response if you want to return another card.
        """
        print("START LINE --------------------------------------------------------")
        room_id = credentials['roomID']
        bearer = credentials['bearer']
        now=datetime.now()
        current_time = now.strftime("%H:%M:%S")
        today=str(date.today())
        print("Script run at", current_time, today)
        print ('ATTACHMENT ACTIONS: ', attachment_actions, ' TYPE: ', type(attachment_actions))
        jsonstr = json.dumps (attachment_actions.__dict__)
        print('TYPE JSONSTR AND JSONSTR ARE: ', type(jsonstr), jsonstr)
        json_actions = json.loads(jsonstr)
        ip = json_actions['_json_data']['inputs']['IP']
        action = json_actions['_json_data']['inputs']['action']
        parsed_url = json_actions['_json_data']['inputs']['expe']
        print('IP IS: ', ip, ' ACTION IS: ', action, ' EXPE IS: ', parsed_url) 
        if ':' in parsed_url: #in this case expe is parsed_url, that is might have the port. In order to connect, the port has to be stripped out.
            expe_list = parsed_url.split(':')
            expe = expe_list[0]
        else:
            expe = parsed_url
        
        activity_list=cdrcheck(parsed_url, credentials[expe][0], credentials[expe][1], ip)
        print(activity_list)
        
        #file_name="Activity_of_" + ip + "_" + today + "_"+ current_time  + ".csv"
        file_name_xlsx="Activity_of_" + ip + "_" + today + "_"+ current_time + ".xlsx"
        file_path=credentials['activity_list_path']
        #complete_name=file_path + file_name
        complete_name_xlsx=file_path + file_name_xlsx
        n=len(activity_list)
        print("ACTIVITY LIST IS MADE OF: " + str(n) + " RECORDS")
        #print(activity_list)
        wb=Workbook()
        ws=wb.active
        #set formatting A1 to B4 bold and size 16 and populate titles and IP info
        columns=['A', 'B']
        for i in range(4):
            for c in columns:
                index=c+str(i+1)
                #print(index)
                if c=='A':
                   c_data=0
                else:
                   c_data=1
                ws[index].font=Font(bold=True, size=16)
                ws[index]=activity_list[c_data][i]
        ws['A6'].font=Font(bold=True)
        ws['B6'].font=Font(bold=True)
        ws['C6'].font=Font(bold=True)
        ws['D6'].font=Font(bold=True)

        rows=len(activity_list)
        cell_converter={'0':'A', '1':'B', '2':'C', '3':'D'}
        for index1 in range(2, rows):
            for index2 in range(4):
              cellindex=cell_converter[str(index2)]+str(index1+4)
              ws[cellindex]=activity_list[index1][index2]
        
        wb.save(complete_name_xlsx)
        

        m = MultipartEncoder({'roomId': room_id,
                      'text': 'Activity of ' + ip,
                      'files': (complete_name_xlsx, open(complete_name_xlsx, 'rb'),
                      'application/vnd.ms-excel')})

        r = requests.post('https://webexapis.com/v1/messages', data=m,
                  headers={'Authorization': 'Bearer ' + bearer,
                  'Content-Type': m.content_type})

        print (r.text)                
        

